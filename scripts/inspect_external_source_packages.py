#!/usr/bin/env python3
"""Inspect external research sources without retaining raw data packages.

The reconnaissance has two independent goals:

1. establish an authoritative, machine-readable OEWS access path and observed schema;
2. test technical access to the published historical RPS replication endpoint.

Technical accessibility is never interpreted as publication or redistribution permission.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import openpyxl
import requests

OEWS_ARCHIVE_URL = "https://www.bls.gov/oes/special-requests/oesm25in4.zip"
OEWS_TEXT_ROOT = "https://download.bls.gov/pub/time.series/oe"
OEWS_API_URL = "https://api.bls.gov/publicAPI/v1/timeseries/data/"
RPS_REPLICATION_URL = (
    "https://services.informs.org/dataset/download.php?doi=mnsc.2025.02523"
)


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _fetch(
    url: str,
    *,
    timeout: int = 90,
    method: str = "GET",
    json_payload: dict[str, Any] | None = None,
) -> tuple[requests.Response | None, str | None]:
    try:
        response = requests.request(
            method,
            url,
            timeout=timeout,
            allow_redirects=True,
            headers={
                "User-Agent": "genai-at-work-research/0.1 source-validation",
                "Accept": "*/*",
            },
            json=json_payload,
        )
        return response, None
    except requests.RequestException as exc:
        return None, f"{type(exc).__name__}: {exc}"


def _safe_headers(response: requests.Response) -> dict[str, str | None]:
    return {
        "content_type": response.headers.get("content-type"),
        "content_length": response.headers.get("content-length"),
        "content_disposition": response.headers.get("content-disposition"),
        "last_modified": response.headers.get("last-modified"),
        "etag": response.headers.get("etag"),
    }


def _transport_record(
    *, url: str, response: requests.Response | None, error: str | None
) -> dict[str, Any]:
    if response is None:
        return {"source_url": url, "status": "request_error", "request_error": error}
    return {
        "source_url": url,
        "status": "http_response_received",
        "http_status": response.status_code,
        "final_url": response.url,
        "headers": _safe_headers(response),
        "response_size_bytes": len(response.content),
        "response_sha256": _sha256(response.content),
    }


def _parse_tsv(payload: bytes) -> list[dict[str, str]]:
    text = payload.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text), delimiter="\t"))


def _series_id(*, industry_code: str, occupation_code: str, datatype: str = "01") -> str:
    """Construct a national, unadjusted OEWS series ID from documented components."""
    if len(industry_code) != 6 or len(occupation_code) != 6 or len(datatype) != 2:
        raise ValueError("OEWS series components have invalid width")
    return f"OEUN0000000{industry_code}{occupation_code}{datatype}"


def inspect_oews(*, generated_at: str, source_build_commit: str) -> dict[str, Any]:
    result: dict[str, Any] = {
        "status": "source_paths_probed",
        "provider": "U.S. Bureau of Labor Statistics",
        "dataset": "May 2025 Occupational Employment and Wage Statistics",
        "generated_at_utc": generated_at,
        "source_build_commit": source_build_commit,
        "raw_retained_in_repository": False,
        "archive_probe": {},
        "text_interface": {},
        "api_probe": {},
    }

    archive_response, archive_error = _fetch(OEWS_ARCHIVE_URL)
    archive_record = _transport_record(
        url=OEWS_ARCHIVE_URL, response=archive_response, error=archive_error
    )
    if (
        archive_response is not None
        and archive_response.status_code == 200
        and zipfile.is_zipfile(io.BytesIO(archive_response.content))
    ):
        archive_record["is_zip_archive"] = True
        entries: list[dict[str, Any]] = []
        with zipfile.ZipFile(io.BytesIO(archive_response.content)) as archive:
            for info in archive.infolist():
                entry: dict[str, Any] = {
                    "path": info.filename,
                    "size_bytes": info.file_size,
                    "compressed_size_bytes": info.compress_size,
                }
                if info.filename.lower().endswith(".xlsx"):
                    workbook_bytes = archive.read(info.filename)
                    workbook = openpyxl.load_workbook(
                        io.BytesIO(workbook_bytes), read_only=True, data_only=True
                    )
                    entry["xlsx_sha256"] = _sha256(workbook_bytes)
                    entry["sheets"] = [
                        {
                            "name": sheet.title,
                            "max_row": sheet.max_row,
                            "max_column": sheet.max_column,
                            "first_six_rows": [
                                [None if value is None else str(value) for value in row]
                                for row in sheet.iter_rows(
                                    min_row=1, max_row=6, values_only=True
                                )
                            ],
                        }
                        for sheet in workbook.worksheets
                    ]
                    workbook.close()
                entries.append(entry)
        archive_record["archive_entries"] = entries
    else:
        archive_record["is_zip_archive"] = False
    result["archive_probe"] = archive_record

    mapping_specs = {
        "industry": "oe.industry",
        "occupation": "oe.occupation",
        "datatype": "oe.datatype",
        "area": "oe.area",
        "documentation": "oe.txt",
    }
    mapping_payloads: dict[str, bytes] = {}
    text_interface: dict[str, Any] = {}
    for key, filename in mapping_specs.items():
        url = f"{OEWS_TEXT_ROOT}/{filename}"
        response, error = _fetch(url)
        record = _transport_record(url=url, response=response, error=error)
        if response is not None and response.status_code == 200:
            mapping_payloads[key] = response.content
            record["first_lines"] = response.text.splitlines()[:8]
        text_interface[key] = record
    result["text_interface"] = text_interface

    industry_rows = (
        _parse_tsv(mapping_payloads["industry"]) if "industry" in mapping_payloads else []
    )
    occupation_rows = (
        _parse_tsv(mapping_payloads["occupation"])
        if "occupation" in mapping_payloads
        else []
    )
    datatype_rows = (
        _parse_tsv(mapping_payloads["datatype"]) if "datatype" in mapping_payloads else []
    )
    area_rows = _parse_tsv(mapping_payloads["area"]) if "area" in mapping_payloads else []

    sector_prefixes = {
        "11",
        "21",
        "22",
        "23",
        "31",
        "42",
        "44",
        "48",
        "51",
        "52",
        "53",
        "54",
        "55",
        "56",
        "61",
        "62",
        "71",
        "72",
        "81",
        "99",
    }
    sector_candidates = [
        row
        for row in industry_rows
        if len(row.get("industry_code", "")) == 6
        and row["industry_code"][:2] in sector_prefixes
        and row["industry_code"].endswith("0000")
    ]
    major_occupations = [
        row
        for row in occupation_rows
        if len(row.get("occupation_code", "")) == 6
        and row["occupation_code"].endswith("0000")
        and row["occupation_code"] != "000000"
    ]
    result["observed_mapping_summary"] = {
        "industry_row_count": len(industry_rows),
        "sector_level_candidates": sector_candidates,
        "occupation_row_count": len(occupation_rows),
        "major_occupation_candidates": major_occupations,
        "datatype_rows": datatype_rows,
        "national_area_rows": [
            row for row in area_rows if row.get("area_code") == "0000000"
        ],
    }

    sector_54 = next(
        (
            row
            for row in industry_rows
            if row.get("industry_name")
            == "Professional, Scientific, and Technical Services"
        ),
        None,
    )
    management = next(
        (
            row
            for row in occupation_rows
            if row.get("occupation_name") == "Management Occupations"
        ),
        None,
    )
    if sector_54 and management:
        probe_series = [
            _series_id(
                industry_code=sector_54["industry_code"],
                occupation_code="000000",
            ),
            _series_id(
                industry_code=sector_54["industry_code"],
                occupation_code=management["occupation_code"],
            ),
        ]
        response, error = _fetch(
            OEWS_API_URL,
            method="POST",
            json_payload={
                "seriesid": probe_series,
                "startyear": "2025",
                "endyear": "2025",
            },
        )
        api_record = _transport_record(
            url=OEWS_API_URL, response=response, error=error
        )
        api_record["requested_series"] = probe_series
        if response is not None and response.status_code == 200:
            try:
                api_record["json"] = response.json()
            except requests.JSONDecodeError:
                api_record["json_decode_error"] = True
        result["api_probe"] = api_record

    return result


def inspect_rps(*, generated_at: str, source_build_commit: str) -> dict[str, Any]:
    response, error = _fetch(RPS_REPLICATION_URL)
    result = _transport_record(
        url=RPS_REPLICATION_URL, response=response, error=error
    )
    result.update(
        {
            "provider_surface": "INFORMS Management Science replication-files endpoint",
            "article_doi": "10.1287/mnsc.2025.02523",
            "generated_at_utc": generated_at,
            "source_build_commit": source_build_commit,
            "raw_retained_in_repository": False,
            "rights_conclusion": (
                "none; technical accessibility does not establish reuse or "
                "redistribution permission"
            ),
        }
    )
    if response is None:
        return result

    is_zip = zipfile.is_zipfile(io.BytesIO(response.content))
    result["is_zip_archive"] = is_zip
    if is_zip:
        with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
            result["archive_entries"] = [
                {
                    "path": info.filename,
                    "size_bytes": info.file_size,
                    "compressed_size_bytes": info.compress_size,
                }
                for info in archive.infolist()
            ]
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args()

    generated_at = datetime.now(UTC).isoformat()
    source_build_commit = os.environ.get("SOURCE_BUILD_COMMIT", "unknown")
    args.output_dir.mkdir(parents=True, exist_ok=True)

    # Keep the probes independent: one blocked source surface must not prevent the
    # other source from being tested and recorded.
    rps = inspect_rps(generated_at=generated_at, source_build_commit=source_build_commit)
    oews = inspect_oews(
        generated_at=generated_at, source_build_commit=source_build_commit
    )

    (args.output_dir / "oews_may_2025_source_schema.json").write_text(
        json.dumps(oews, indent=2, sort_keys=True) + "\n"
    )
    (args.output_dir / "rps_replication_endpoint_probe.json").write_text(
        json.dumps(rps, indent=2, sort_keys=True) + "\n"
    )
    (args.output_dir / "README.md").write_text(
        "# Source-unlock reconnaissance - 2026-09-01\n\n"
        "OEWS is probed through both the convenience ZIP surface and BLS's official "
        "time-series mapping/API interfaces. A convenience-download transport block "
        "does not invalidate an independently functioning official interface.\n\n"
        "The INFORMS historical RPS replication-files endpoint is probed independently. "
        "Its result establishes technical accessibility only. No reuse, storage, or "
        "redistribution rights are inferred from a successful response. Raw source "
        "packages are not retained in the repository.\n"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
